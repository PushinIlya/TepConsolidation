#!/usr/bin/env python
# coding: utf-8

# In[ ]:


'''
Скрипт для консолидации выгруженных из 1С ТЭП и их распределения по папкам 
в соответствии с периодом, по которому они сформированы
'''
# импортируем библиотеки и модули
import numpy as np
import pandas as pd
import os
import re
import math
import shutil
import copy
from itertools import groupby, chain
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


# In[ ]:


# запишем путь к папке с исходными данными в переменную
source = os.path.join('C:',
                      'Исходные данные')

# создадим список с названиями всех файлов в папке с исходными данными с расширением .xlsx
raw_data_name_list = [raw_data for raw_data in os.listdir(source) if raw_data.startswith('ТЭП') and
                      raw_data.endswith('.xlsx')]

print(f'Исходные данные: {raw_data_name_list}')


# In[ ]:


print('Выполнение скрипта "tep_consolidation" (консолидация ТЭП)')
print(f'Чтение ТЭП из "{source}"')

# создадим справочник из всех датафреймов в списке 'raw_data_name_list'
all_raw_data = {raw_data:pd.read_excel(os.path.join(source, f'{raw_data}')) for raw_data in tqdm(raw_data_name_list)}


# In[ ]:


# удалим столбцы, содержащие только пустые значения
all_raw_data = {raw_data[0]: raw_data[1].dropna(axis=1, how='all', inplace=False) for raw_data in all_raw_data.items()}


# In[ ]:


# в первую строку первого столбца скопируем название первого столбца
for key, value in all_raw_data.items():
    value.iloc[0,0] = value.columns[0]


# In[ ]:


# переименуем столбцы в каждом датафрейме
for raw_data in all_raw_data.values():
    raw_data.columns = list(range(0,10))


# In[ ]:


# разделим столбцы с суммами на 1000
for raw_data in all_raw_data.values():
    raw_data.iloc[4:, [2,3,4,6,7,8]] = raw_data.iloc[4:, [2,3,4,6,7,8]] / 1000


# In[ ]:


# создадим список 'source_file_path' и заполним его значениями путей к исходным файлам
source_file_path = []
for name in raw_data_name_list:
    source_file_path.append(os.path.join(source, name))


# In[ ]:


# объединим списки 'raw_data_name_list' и 'source_file_path'
# в список кортежей с названием файла и путём к нему
source_file = list(zip(raw_data_name_list, source_file_path))


# In[ ]:


'''
Цикл, который перезаписывает каждый файл из 'all_raw_data', чтобы сохранить внесённые ранее изменения в его структуру,
а также сохраняет файл в формате xlsx в первоначальную папку
'''
print(f'Перезапись ТЭП из "{source}"')
for raw_data in tqdm(all_raw_data.items()):
    for file in source_file:
        if raw_data[0] == file[0]:
            raw_data[1].to_excel(file[1], header=False, index=False)


# In[ ]:


# создадим список с периодами формирования ТЭП для каждого файла
period_list = []
for raw_data in all_raw_data.values():
    period_list.append(raw_data.iloc[1,2])


# In[ ]:


def extract_month_year(text):
    '''
    Функция для извлечения месяца и года из строки
    \w+ - группа захвата символов, которые состоят из букв, цифр или символа подчёркивания и могут повторяться один или более раз
    \s - указывает на наличие пробела
    \d{4}) - группа захвата, представляющая четырехзначное число
    '''
    pattern = r'(\w+\s\d{4})'
    match = re.search(pattern, text)
    if match:
        return match.group(1)
    else:
        return None


# In[ ]:


# применим функцию 'extract_month_year'
period_list = [extract_month_year(period) for period in period_list]


# In[ ]:


# переименуем каждый датафрейм, оставив только название компании в имени
all_raw_data = {raw_data[0][raw_data[0].rfind('_')+1:]: raw_data[1] for raw_data in all_raw_data.items()}


# In[ ]:


# удалим строки, содержащие только пустые значения
all_raw_data = {raw_data[0]: raw_data[1].dropna(axis=0, how='all', inplace=False) for raw_data in all_raw_data.items()}


# In[ ]:


# функция для извлечения номера месяца по названию месяца
def month_number(month_name):
    months = {'январь': '01',
              'февраль': '02', 
              'март': '03',
              'апрель': '04', 
              'май': '05', 
              'июнь': '06',
              'июль': '07',
              'август': '08',
              'сентябрь': '09',
              'октябрь': '10', 
              'ноябрь': '11', 
              'декабрь': '12'}
    month_number = months[month_name]
    return month_number


# In[ ]:


# удалим лишние строки в датафреймах
all_raw_data = {raw_data[0]: raw_data[1][4:].reset_index(drop=True) for raw_data in all_raw_data.items()}


# In[ ]:


# словарь с именами столбцов и типами данных
type_of_columns = {2: 'float64', 
                   3: 'float64', 
                   4: 'float64',
                   5: 'float64', 
                   6: 'float64', 
                   7: 'float64', 
                   8: 'float64',
                   9: 'float64'}


# In[ ]:


# изменим тип данных в каждом датафрейме в соответствии со словарём 'type_of_columns'
all_raw_data = {raw_data[0]: raw_data[1].astype(type_of_columns) for raw_data in all_raw_data.items()}


# In[ ]:


# заполним значения в первом столбце вниз
for key, value in all_raw_data.items():
    value[0] = value[0].fillna(method='ffill', axis=0)


# In[ ]:


# создадим списоки, куда добавим новые кодификаторы
codificators_index = []
codificators_values = []


# In[ ]:


# цикл для заполнения списков с новыми кодификаторами
for key, value in all_raw_data.items():
    for index, element in enumerate(list(value[0])):
        if list(value[0]).count(element) > 1:
            codificators_index.append(index)
            codificators_values.append(element)


# In[ ]:


# разделим список 'codificator_values' на группы из повторяющихся значений
codificators_values_groups = []
for key, group in groupby(codificators_values):
    codificators_values_groups.append(list(group))


# In[ ]:


# с помощью циклов создадим новые кодификаторы
for i in range(len(codificators_values_groups)):
    for j in range(len(codificators_values_groups[i])):
        codificators_values_groups[i][j] = codificators_values_groups[i][j] + str(j-1+1)+str('.')


# In[ ]:


# преобразуем список списков 'codificators_values_groups' в список
codificators_values_groups = list(chain.from_iterable(codificators_values_groups))


# In[ ]:


# создадим словарь с индексами и значениями новых кодификаторов
codificators = dict(zip(codificators_index, codificators_values_groups))


# In[ ]:


# заполним столбец с № п/п новыми кодификаторами
# пройдёмся циклом по каждой таблице в 'all_raw_data'
for key, value in all_raw_data.items():
    # пройдёмся циклом по каждому ключу в 'codificators'
    for replace_index, replace_value in codificators.items():
        # заменим значение в столбце '0' для строки с заданным индексом
        value.loc[replace_index, 0] = replace_value


# In[ ]:


# заменим '.0' на '' в столбце '0' в каждой таблице
for key, value in all_raw_data.items():
    value[0] = value[0].replace('\.0', '', regex=True)


# In[ ]:


new_columns = ['№ п/п',
               'Наименование показателя',
               f'План ({period_list[0]}), тыс.руб.',
               f'Факт ({period_list[0]}), тыс.руб.',
               f'Отклонение ({period_list[0]}), тыс.руб.',
               f'Отклонение ({period_list[0]}), %',
               'План (Итого с начала года), тыс.руб.',
               'Факт (Итого с начала года), тыс.руб.',
               'Отклонение (Итого с начала года), тыс.руб.',
               'Отклонение (Итого с начала года), %']
# переименуем столбцы в каждой таблице
for key, value in all_raw_data.items():
    value.columns = new_columns


# In[ ]:


# создадим столбец 'sort_column' для сортировки таблицы 'consolidated_data'
for key, value in all_raw_data.items():
    value['sort_column'] = value.index


# In[ ]:


# объединим данные по каждому датафрейму в единый датафрейм, путём суммирования построчно
consolidated_data = (pd.concat(all_raw_data.values())
                     .groupby(['№ п/п', 'Наименование показателя', 'sort_column'])
                     .sum()
                     .sort_values(['№ п/п'])
                     .reset_index()
                    )


# In[ ]:


# отсортируем таблицу 'consolidated_data'
consolidated_data = consolidated_data.sort_values(by='sort_column').reset_index(drop=True)


# In[ ]:


# удалим столбец 'sort_column' в каждой таблице
for key, value in all_raw_data.items():
    value.drop('sort_column', axis=1, inplace=True)


# In[ ]:


# удалим столбец 'sort_column'
consolidated_data.drop('sort_column', axis=1, inplace=True)


# In[ ]:


# рассчитаем значения в столбце 'Отклонение (отн., %)' за месяц
# рассчитаем значения в столбце 'Отклонение (отн., %)' за весь период
consolidated_data[f'Отклонение ({period_list[0]}), %'] = consolidated_data[f'Отклонение ({period_list[0]}), тыс.руб.']/abs(consolidated_data[f'План ({period_list[0]}), тыс.руб.'])*100
consolidated_data['Отклонение (Итого с начала года), %'] = consolidated_data['Отклонение (Итого с начала года), тыс.руб.']/abs(consolidated_data['План (Итого с начала года), тыс.руб.'])*100


# In[ ]:


# при расчёте значений 'Отклонение (отн., %)' при делении на 0 получились бесконечные числа, заменим их на 100
consolidated_data.replace([np.inf, -np.inf], 100, inplace=True)


# In[ ]:


# создадим excel файл 'Сводный_ТЭП' и добавим в него все датафреймы с исходными данными
with pd.ExcelWriter(os.path.join(source, f'Сводный_ТЭП_{period_list[0]}.xlsx')) as source_file:
    print(f'Создание листов excel из ТЭП')
    for raw_data in tqdm(all_raw_data.items()):
        raw_data[1].to_excel(source_file,
                             sheet_name=raw_data[0][raw_data[0].rfind('_')+1:].replace('.xlsx', ''),
                             header=True,
                             index=False)


# In[ ]:


book_path = os.path.join(source, f'Сводный_ТЭП_{period_list[0]}.xlsx') # создадим путь к итоговому excel файлу
book = load_workbook(book_path) # загрузим excel файл для редактирования


# In[ ]:


columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'] # создадим список столбцов, к которым будем обращаться


# In[ ]:


# создадим новый лист
new_sheet = book.create_sheet('Консолидация')

# загрузим объединённый датафрейм на новый лист
for row in dataframe_to_rows(consolidated_data, index=False, header=True):
    new_sheet.append(row)


# In[ ]:


# создадим таблицы на каждом листе
for sheet in book:
    if sheet.title == 'Консолидация':
        table = Table(displayName=sheet.title, ref=f'A1:J{sheet.max_row}') # создадим таблицу
        style = TableStyleInfo(name='TableStyleLight13', # создадим стиль таблицы
                               showFirstColumn=False,
                               showLastColumn=False,
                               showRowStripes=True,
                               showColumnStripes=True)
        table.tableStyleInfo = style # применим стиль
        sheet.add_table(table) # добавим таблицу на лист
    else:
        table = Table(displayName=sheet.title, ref=f'A1:J{sheet.max_row}')
        style = TableStyleInfo(name='TableStyleLight13',
                               showFirstColumn=False,
                               showLastColumn=False,
                               showRowStripes=True,
                               showColumnStripes=True)
        table.tableStyleInfo = style
        sheet.add_table(table)


# In[ ]:


# функция для группировки строк на каждом листе, кроме 'Консолидация' в excel файле на основе столбца '№ п/п'
def rows_grouping(lvl):
    # группировка 1-го уровня
    # создадим список с индексами строк для группировки 1-го уровня
    rows_start = {raw_data[0]: list(raw_data[1]['№ п/п']
                                    .str.count('\\.')[lambda x: x == 1]
                                    .index) 
                  for raw_data in all_raw_data.items()}
        
    # добавим в список с индексами строк для группировки 1-го уровня последнюю строку на листе,
    # чтобы группировка была корректной 
    for raw_data in all_raw_data.items():
        rows_start[raw_data[0]].append(list(raw_data[1]['№ п/п'].index)[-1])
            
    # создадим аналогичный список с индексами строк для группировки 1-го уровня,
    # только смещённый на 1-ну позицию
    rows_end = {raw_data[0]: rows_start[raw_data[0]][1:]
                for raw_data in all_raw_data.items()}
    # пересоздадим список с индексами строк для группировки 1-го уровня, так как до этого его изменили
    rows_start = {raw_data[0]: list(raw_data[1]['№ п/п']
                                    .str.count('\\.')[lambda x: x == 1]
                                    .index)
                  for raw_data in all_raw_data.items()}
    # объединим оба списка в список кортежей
    rows = {raw_data[0]: list(zip(rows_start[raw_data[0]], rows_end[raw_data[0]]))
            for raw_data in all_raw_data.items()}

    return rows


# In[ ]:


# функция для группировки строк в excel файле на основе столбца '№ п/п' на листе 'Консолидация'
# код аналогичен функции 'rows_grouping'
def rows_grouping_conso(lvl):
    if lvl == 1:
        rows_start = list(consolidated_data['№ п/п'].str.count('\\.')[lambda x: x == 1].index)
        rows_start.append(list(consolidated_data['№ п/п'].index)[-1])
    
        rows_end = rows_start[1:]
        rows_start = list(consolidated_data['№ п/п'].str.count('\\.')[lambda x: x == 1].index)
        rows = list(zip(rows_start, rows_end))
        
        return rows


# In[ ]:


# создадим для каждого уровня группировки на каждом листе, кроме 'Консолидация' список кортежей
rows_1 = rows_grouping(1)
# создадим для каждого уровня группировки на листе 'Консолидация' список кортежей
rows_1_conso = rows_grouping_conso(1)


# In[ ]:


# цикл для изменения ширины столбцов на каждом листе в excel файле
for sheet in book:
    for column in columns:
        sheet.column_dimensions[column].width = 20
    sheet.column_dimensions['B'].width = 65


# In[ ]:


# цикл для изменения формата ячеек на каждом листе в excel файле
for sheet in book:
    for column in columns:
        if column in ['A', 'B']:
            for cell in sheet[column:column]:
                sheet[cell.coordinate].number_format = BUILTIN_FORMATS[0]  
        else:
            for cell in sheet[column:column]:
                sheet[cell.coordinate].number_format = BUILTIN_FORMATS[3]


# In[ ]:


# цикл для изменения стиля и выравнивания ячеек на каждом листе в excel файле
for sheet in book:
    if sheet.title == 'Консолидация':
        for column in columns:
            # изменение параметров шрифта
            sheet[f'{column}1'].font = Font(bold=True, 
                                            color='FFFFFFFF')
            # изменение выравнивания
            sheet[f'{column}1'].alignment = Alignment(horizontal='center',
                                                      vertical='center',
                                                      wrap_text=True)
            
            # изменение заливки и шрифта сгруппированных строк
            for row in list(consolidated_data['№ п/п'][lambda x: x.str.count('\\.') == 1].index):
                pattern_fill = PatternFill(fill_type='solid',
                                           start_color='DAEEF3',
                                           end_color='DAEEF3')
                sheet[f'{column}{row+2}'].fill = pattern_fill # к строке добавляем 2 из-за смещения нумерации
                sheet[f'{column}{row+2}'].font = Font(bold=True)
        
        # изменение выравнивания числовых ячеек
        for cell_tuple in sheet[f'C2:J{sheet.max_row}']:
            for cell in cell_tuple:
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center')
    else:
        for column in columns:
            # изменение параметров шрифта
            sheet[f'{column}1'].font = Font(bold=True, 
                                            color='FFFFFFFF')
            # изменение выравнивания
            sheet[f'{column}1'].alignment = Alignment(horizontal='center',
                                                      vertical='center',
                                                      wrap_text=True)
            
        # изменение заливки и шрифта сгруппированных строк
        for raw_data in all_raw_data.items():
            rows_1 = list(raw_data[1]['№ п/п'].str.count('\\.')[lambda x: x == 1].index)
            if raw_data[0].replace('.xlsx', '') == sheet.title:
                for column in columns:
                    for row in rows_1:
                        patern_fill = PatternFill(fill_type='solid',
                                                  start_color='DAEEF3',
                                                  end_color='DAEEF3')
                        sheet[f'{column}{row+2}'].fill = patern_fill # к строке добавляем 1 из-за смещения нумерации
                        sheet[f'{column}{row+2}'].font = Font(bold=True)

        # изменение выравнивания числовых ячеек
        for cell_tuple in sheet[f'C2:J{sheet.max_row}']:
            for cell in cell_tuple:
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center')


# In[ ]:


# цикл для изменения отступа ячеек на каждом листе в excel файле
for sheet in book:
    if sheet.title == 'Консолидация':
        rows_1 = list(consolidated_data['№ п/п'].str.count('\\.')[lambda x: x == 2].index)
        for row in rows_1:
            sheet[f'B{row+2}'].alignment = Alignment(indent=1) # к строке добавляем 2 из-за смещения нумерации
    else:
        rows_1 = list(all_raw_data[f'{sheet.title}.xlsx']['№ п/п'].str.count('\\.')[lambda x: x == 2].index)
        for row in rows_1:
            sheet[f'B{row+2}'].alignment = Alignment(indent=1) # к строке добавляем 2 из-за смещения нумерации


# In[ ]:


book.save(book_path) # сохраним excel файл


# In[ ]:


# создадим список с расположением новых папок
folder_location = os.path.join(source,
                               period_list[0][-4:],
                               f'{month_number(str.lower(period_list[0][:-5]))}_{str.lower(period_list[0][:-5])}')
# создаются только те папки, которые до этого не существовали
if not os.path.exists(folder_location):
    os.makedirs(folder_location)


# In[ ]:


print(f'Распределение ТЭП и итогового файла из "{source}" по папкам')
# скопируем итоговый файл с расширением .xlsx в новую папку
shutil.copy(book_path, folder_location)
# удалим итоговый файл с расширением .xlsx из первоначальной папки
os.remove(book_path)

# цикл, который копирует каждый файл из 'raw_data_list' в новую папку и удаляет его из папки c исходными данными
for name in tqdm(raw_data_name_list):
    # скопируем исходные данные с расширением .xlsx в новую папку
    shutil.copy(os.path.join(source, name), folder_location)
    # удалим исходные данные с расширением .xlsx из первоначальной папки
    os.remove(os.path.join(source, name))

